#!/usr/bin/env python3
"""
build_lab_data.py — generate the mock-data workbooks each lab needs.

Every workbook is written with a real Excel TABLE (not a loose range), because
Power Apps and Power Automate can only bind to a named table — the single most
common beginner failure in this course, and a point the labs make explicitly.
"""
import os, random, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
REPO = os.path.dirname(HERE)
LABS = os.path.join(REPO, "labs")

random.seed(20260904)   # deterministic mock data


def write_table(path, sheet_name, table_name, headers, rows):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    for r in rows:
        ws.append(r)
    last_col = get_column_letter(len(headers))
    ref = f"A1:{last_col}{len(rows) + 1}"
    t = Table(displayName=table_name, ref=ref)
    t.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(t)
    for i, h in enumerate(headers, start=1):
        width = max(len(str(h)), *(len(str(r[i-1])) for r in rows)) + 2 if rows else len(h) + 2
        ws.column_dimensions[get_column_letter(i)].width = min(width, 42)
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path


# --------------------------------------------------------------- service calls
INSTALL = ["Solar Panel", "Wind Turbine", "Battery Wall", "Inverter", "EV Charger"]
PROBLEMS = [
    "Unit not producing expected output",
    "Intermittent fault light on controller",
    "Loud noise during operation",
    "App shows device offline",
    "Firmware update failed",
    "Water ingress in junction box",
    "Breaker trips under load",
    "Display panel unresponsive",
]
ENGINEERS = ["Marcus Tan", "Priya Nair", "Wei Ling Ho", "Daniel Lim", "Aisha Rahman"]
REPORTERS = ["ops@kineteco.example", "helpdesk@kineteco.example", "field@kineteco.example"]
RESOLUTIONS = ["Replaced faulty component", "Reseated connector and retested",
               "Firmware reflashed", "Escalated to manufacturer", ""]


def service_calls(n=40):
    rows = []
    base = datetime.date(2026, 8, 1)
    for i in range(n):
        d = base + datetime.timedelta(days=random.randint(0, 30))
        resolved = random.random() > 0.35
        urgent = random.random() > 0.75
        rows.append([
            random.choice(INSTALL),
            random.choice(REPORTERS),
            d.strftime("%Y-%m-%d"),
            ("URGENT: " if urgent else "") + random.choice(PROBLEMS),
            random.choice(ENGINEERS),
            random.choice(RESOLUTIONS[:-1]) if resolved else "",
            (d + datetime.timedelta(days=random.randint(1, 6))).strftime("%Y-%m-%d") if resolved else "",
        ])
    return rows


# --------------------------------------------------------------- leave log
STAFF = [
    ("Marcus Tan", "marcus.tan@kineteco.example"),
    ("Priya Nair", "priya.nair@kineteco.example"),
    ("Wei Ling Ho", "weiling.ho@kineteco.example"),
    ("Daniel Lim", "daniel.lim@kineteco.example"),
    ("Aisha Rahman", "aisha.rahman@kineteco.example"),
    ("Joseph Koh", "joseph.koh@kineteco.example"),
]
LEAVE_TYPES = ["Annual", "Medical", "Childcare", "Unpaid"]
STATUSES = ["Approved", "Rejected", "Pending", "Submitted"]


def leave_log(n=25):
    rows = []
    base = datetime.date(2026, 7, 1)
    for i in range(n):
        name, email = random.choice(STAFF)
        start = base + datetime.timedelta(days=random.randint(0, 70))
        days = random.choice([1, 1, 2, 3, 5])
        rows.append([
            name, email, random.choice(LEAVE_TYPES),
            start.strftime("%Y-%m-%d"), days,
            random.choice(["Family matters", "Medical appointment", "Vacation",
                           "Childcare", "Personal"]),
            random.choice(STATUSES),
            random.choice(["", "", "Approved as requested", "Team coverage needed"]),
        ])
    return rows


# --------------------------------------------------------------- survey
def survey(n=30):
    rows = []
    base = datetime.date(2026, 8, 1)
    for i in range(n):
        d = base + datetime.timedelta(days=random.randint(0, 30))
        rows.append([
            d.strftime("%Y-%m-%d"),
            random.choice(["Engineering", "Operations", "Sales", "Support", "Finance"]),
            random.randint(1, 5), random.randint(1, 5), random.randint(1, 5),
            random.choice(["Yes", "No"]),
            random.choice(["Good tools and support", "Workload is high",
                           "Great team culture", "More training needed", ""]),
        ])
    return rows


def main():
    made = []

    # Lab 1 & 2 — the manual process the learner scans for opportunities
    for lab in ("lab-01", "lab-02"):
        made.append(write_table(
            os.path.join(LABS, lab, "data", "KinetEco Service Calls.xlsx"),
            "Calls", "ServiceCalls",
            ["Install Type", "Reported By", "Date Reported", "Problem",
             "Assigned to", "Resolution", "Date Resolved"],
            service_calls()))

    # Labs 3-5, 7, 10 — the service-call table the flows and the first app use
    for lab in ("lab-03", "lab-04", "lab-05", "lab-07", "lab-10"):
        made.append(write_table(
            os.path.join(LABS, lab, "data", "KinetEco Service Calls.xlsx"),
            "Calls", "ServiceCalls",
            ["Install Type", "Reported By", "Date Reported", "Problem",
             "Assigned to", "Resolution", "Date Resolved"],
            service_calls()))

    # Labs 6, 11, 12, 13, 14 — the leave log the approval chain writes to
    for lab in ("lab-06", "lab-11", "lab-12", "lab-13", "lab-14"):
        made.append(write_table(
            os.path.join(LABS, lab, "data", "LeaveLog.xlsx"),
            "Leave", "LeaveLog",
            ["Applicant", "Email", "LeaveType", "StartDate", "Days",
             "Reason", "Status", "Comments"],
            leave_log()))

    # Lab 8 & 9 — a survey set for the blank-canvas app and the API lab
    for lab in ("lab-08", "lab-09"):
        made.append(write_table(
            os.path.join(LABS, lab, "data", "Employee Survey.xlsx"),
            "Survey", "SurveyResponses",
            ["Date", "Department", "Satisfaction", "Workload", "Support",
             "Recommend", "Comments"],
            survey()))

    print(f"built {len(made)} workbook(s)")
    for m in made:
        print("  ", os.path.relpath(m, REPO))


if __name__ == "__main__":
    main()
